"""Parse the H1B Excel sheet into a JSON file the scrapers consume."""
import json
from pathlib import Path

import openpyxl

from config import COMPANIES_JSON, EXCEL_PATH, EXTRA_COMPANIES_JSON, EXCLUDED_COMPANIES


def _filter_excluded(companies: list[dict]) -> list[dict]:
    return [c for c in companies if (c.get("name") or "").strip() not in EXCLUDED_COMPANIES]


def load_from_excel(path: Path = EXCEL_PATH) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    companies: list[dict] = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "summary":
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        for r in rows[1:]:
            if not r or not r[1]:
                continue
            companies.append({
                "category": sheet_name,
                "name": r[1],
                "industry": r[2],
                "common_roles": r[3],
                "url": r[4],
                "h1b_level": r[5],
                "notes": r[6],
            })
    return companies


def main() -> None:
    companies = _filter_excluded(load_from_excel())
    COMPANIES_JSON.write_text(json.dumps(companies, indent=2))
    if not EXTRA_COMPANIES_JSON.exists():
        EXTRA_COMPANIES_JSON.write_text("[]\n")
    print(f"Wrote {len(companies)} companies → {COMPANIES_JSON}")


if __name__ == "__main__":
    main()
